from django.http import HttpResponse
from django.contrib.admin.views.decorators import staff_member_required
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Pedido, Cliente


@staff_member_required
def exportar_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=pedidos.pdf'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph('Reporte de Pedidos', styles['Title']))

    pedidos = Pedido.objects.select_related('cliente').prefetch_related('detalles__producto').all()
    datos = [['ID', 'Cliente', 'Fecha', 'Estado', 'Total']]
    total_general = 0

    for p in pedidos:
        total_pedido = sum(d.cantidad * d.producto.precio for d in p.detalles.all())
        total_general += total_pedido
        datos.append([
            str(p.id),
            p.cliente.nombre,
            p.fecha.strftime('%Y-%m-%d'),
            p.estado,
            f'${total_pedido:.2f}',
        ])

    # Fila de total general
    datos.append(['', '', '', 'TOTAL GENERAL:', f'${total_general:.2f}'])

    col_count = len(datos[0])
    tabla = Table(datos, colWidths=[40, 140, 90, 90, 80])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#EBF5FB')]),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
        # Estilo para la fila de total general
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#D5E8D4')),
        ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN',      (col_count-2, -1), (col_count-1, -1), 'RIGHT'),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    return response


@staff_member_required
def exportar_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    azul_fill = PatternFill('solid', fgColor='1E3A5F')
    blanco_font = Font(bold=True, color='FFFFFF')
    verde_fill = PatternFill('solid', fgColor='D5E8D4')
    negrita_font = Font(bold=True)

    encabezados = ['ID', 'Cliente', 'Fecha', 'Estado', 'Total']
    for col, enc in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=col, value=enc)
        c.fill = azul_fill
        c.font = blanco_font
        c.alignment = Alignment(horizontal='center')

    pedidos = Pedido.objects.select_related('cliente').prefetch_related('detalles__producto').all()
    total_general = 0

    for fila, p in enumerate(pedidos, 2):
        total_pedido = sum(d.cantidad * d.producto.precio for d in p.detalles.all())
        total_general += total_pedido
        ws.cell(row=fila, column=1, value=p.id)
        ws.cell(row=fila, column=2, value=p.cliente.nombre)
        ws.cell(row=fila, column=3, value=p.fecha.strftime('%Y-%m-%d'))
        ws.cell(row=fila, column=4, value=p.estado)
        ws.cell(row=fila, column=5, value=round(total_pedido, 2))

    # Fila de total general
    ultima_fila = len(list(pedidos)) + 2
    ws.cell(row=ultima_fila + 1, column=4, value='TOTAL GENERAL:').font = negrita_font
    ws.cell(row=ultima_fila + 1, column=5, value=round(total_general, 2)).font = negrita_font
    for col in range(1, 6):
        ws.cell(row=ultima_fila + 1, column=col).fill = verde_fill

    wb.save(response)
    return response


@staff_member_required
def exportar_clientes_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=clientes.pdf'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph('Reporte de Clientes', styles['Title']))

    datos = [['ID', 'Nombre', 'Correo', 'Teléfono']]
    for c in Cliente.objects.all():
        datos.append([str(c.id), c.nombre, c.correo, c.telefono or ''])

    tabla = Table(datos, colWidths=[40, 160, 180, 100])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#EBF5FB')]),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    elementos.append(tabla)
    doc.build(elementos)
    return response


@staff_member_required
def exportar_clientes_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clientes'

    azul_fill = PatternFill('solid', fgColor='1E3A5F')
    blanco_font = Font(bold=True, color='FFFFFF')

    encabezados = ['ID', 'Nombre', 'Correo', 'Teléfono']
    for col, enc in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=col, value=enc)
        c.fill = azul_fill
        c.font = blanco_font
        c.alignment = Alignment(horizontal='center')

    for fila, cliente in enumerate(Cliente.objects.all(), 2):
        ws.cell(row=fila, column=1, value=cliente.id)
        ws.cell(row=fila, column=2, value=cliente.nombre)
        ws.cell(row=fila, column=3, value=cliente.correo)
        ws.cell(row=fila, column=4, value=cliente.telefono or '')

    wb.save(response)
    return response
